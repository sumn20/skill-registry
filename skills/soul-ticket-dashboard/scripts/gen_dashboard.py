#!/usr/bin/env python3
"""Soul 工单仪表盘 URL 自动拼接 + 排查结论脚本

功能:
  1. 读取工单 Excel，拼接 TRTC 监控搜索页 URL (U列)
  2. 通过 Playwright 持久化浏览器上下文（自动 OA 登录），调用 getRoomList API
  3. 获取 CommId 并拼接通话详情页 URL (V列/22列)
  4. 根据反馈文本 + 仪表盘信息生成初步排查结论 (T列)

用法:
  # 仅拼接搜索页 URL (无需登录)
  python3 gen_dashboard.py <input.xlsx> <output.xlsx>

  # 获取通话详情页 (首次需要 iOA 验证登录, 之后自动复用)
  python3 gen_dashboard.py <input.xlsx> <output.xlsx> --detail

  # 同时生成排查结论
  python3 gen_dashboard.py <input.xlsx> <output.xlsx> --detail --analyze
"""
import sys
import os
import json
import re
import argparse
import time
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

# ─── 常量 ───────────────────────────────────────────────
SDKAPPID_MAP = {
    '语音匹配': '1600050511',
    '私聊语音': '1600050509',
    '视频匹配': '1600050511',
}

LINK_FONT = Font(color='0563C1', underline='single')
TAG_COL = 19          # S列: 问题标签 (自动填充)
CONCLUSION_COL = 20   # T列: 排查结论
SEARCH_COL = 21       # U列: 仪表盘 URL
DETAIL_COL = 22       # V列: 通话详情 URL
LOG_COL = 23          # W列: 日志 (已有)
FOLLOW_COL = 24       # X列: 跟进人 (已有)
NOTE_COL = 25         # Y列: 备注 (已有)
TIME_RANGE = 1800     # ±30分钟

# Playwright 持久化上下文目录 (保存登录 cookie)
PW_USER_DATA_DIR = os.path.expanduser('~/.soul-trtc-profile')

# OA 登录相关
OA_LOGIN_HOST = 'std.passport.woa.com'
TRTC_MONITOR_HOST = 'trtc-monitor.woa.com'
TRTC_API_HOST = 'avmonitor.trtc.woa.com'


# ─── 列定位 ─────────────────────────────────────────────
def find_col(header, name):
    for i, h in enumerate(header):
        if h and str(h).strip() == name:
            return i + 1
    return None


# ─── URL 构造 ────────────────────────────────────────────
def build_search_url(sdkappid, channel_id, start_ts, end_ts):
    return (
        f"https://trtc-monitor.woa.com/"
        f"?sdkAppId={sdkappid}"
        f"&roomNum={channel_id}"
        f"&environment=inland"
        f"&startTs={start_ts}"
        f"&endTs={end_ts}"
    )


def build_roomlist_api_url(sdkappid, channel_id, start_ts, end_ts):
    return (
        f"https://avmonitor.trtc.woa.com/getRoomList"
        f"?RoomNum={channel_id}"
        f"&SourceType=0"
        f"&CallStatus=0"
        f"&UserId="
        f"&SdkAppId={sdkappid}"
        f"&PageIndex=0"
        f"&PageSize=10"
        f"&StartTs={start_ts}"
        f"&EndTs={end_ts}"
    )


def build_detail_url(comm_id, room_num, room_str, create_time,
                     destroy_time, duration, finished, sdkappid):
    return (
        f"https://trtc-monitor.woa.com/trtc/monitor/call-details"
        f"?commId={comm_id}"
        f"&userId="
        f"&roomNum={room_num}"
        f"&roomStr={room_str}"
        f"&createTime={create_time}"
        f"&destroyTime={destroy_time}"
        f"&duration={duration}"
        f"&finished={'true' if finished else 'false'}"
        f"&sdkAppid={sdkappid}"
        f"&environment=inland"
    )


def to_timestamp(feedback_time):
    if isinstance(feedback_time, datetime):
        return int(feedback_time.timestamp())
    try:
        return int(datetime.strptime(
            str(feedback_time), '%Y-%m-%d %H:%M:%S'
        ).timestamp())
    except (ValueError, TypeError):
        return None


# ─── 解析反馈文本 ────────────────────────────────────────
def parse_feedback_text(text):
    """从反馈文本中提取结构化信息"""
    if not text:
        return {}
    text = str(text)
    info = {}

    # 问题描述 (# 之前的部分)
    desc_match = re.match(r'^(.+?)(?:\s*#|$)', text)
    if desc_match:
        info['description'] = desc_match.group(1).strip()

    # 问题类型
    m = re.search(r'#问题类型:([^ ,，]+)', text)
    if m:
        info['problem_type'] = m.group(1)

    # 时间点
    m = re.search(r'时间点:([^ ,，]+)', text)
    if m:
        info['time_point'] = m.group(1)

    # 场景
    m = re.search(r'scene:([^ ,，]+)', text)
    if m:
        info['scene'] = m.group(1)

    # 扬声器
    m = re.search(r'isSpeakerEnable:(true|false)', text)
    if m:
        info['speaker_enabled'] = m.group(1) == 'true'

    # 耳机类型
    m = re.search(r'headphoneType:([^ ,，]+)', text)
    if m:
        info['headphone_type'] = m.group(1)

    return info


# ─── TRTC 事件码映射 ─────────────────────────────────────
EVENT_CODES = {
    5001: '开始进房', 5002: '获取IP成功', 5003: '进房成功',
    5009: '发送首帧音频', 5010: 'UDP进房',
    7000: '进房(后台)', 7001: '正常退出(后台)',
    3001: '采集音频', 3002: '开始播放', 3004: '扬声器播放',
    3005: '采集打断开始', 3006: '采集打断恢复',
    3007: '音频接口切换', 3008: '采集硬件参数', 3009: '播放硬件参数',
    3010: '音频编码', 3011: '本地静音',
    3012: '采集启动', 3013: '采集启动失败',
    3014: '麦克风权限',
    2001: '前后台切换',
    4006: '摄像头操作', 4007: '视频编码', 4008: '视频解码',
    6001: 'startLocalAudio', 6002: 'stopLocalAudio',
}


def _analyze_events(events):
    """分析事件列表，返回结构化分析结果 (flags + findings)
    
    返回 dict:
      - findings: list[str]  事件描述列表
      - flags: dict          结构化标志位
        - enter_room_ok: bool
        - first_audio: bool
        - exit_normal: bool
        - has_bg_switch: bool     是否有前后台切换
        - bg_no_return: bool      切后台后未回来
        - has_mute: bool          是否有mute操作
        - has_capture_interrupt: bool  是否有采集打断
        - capture_interrupt_recovered: bool  采集打断是否恢复
        - has_start_local_audio: bool  是否调用了startLocalAudio
        - has_stop_local_audio: bool
        - capture_start_failed: bool   采集启动失败
        - no_mic_permission: bool      无麦克风权限
        - bg_capture_silent: bool      后台采集无声(切后台后未恢复采集)
        - timeout_exit: bool           超时退房
        - sample_rate_changed: bool    采样率变更
    """
    findings = []
    flags = {
        'enter_room_ok': False, 'first_audio': False, 'exit_normal': False,
        'has_bg_switch': False, 'bg_no_return': False,
        'has_mute': False, 'has_capture_interrupt': False,
        'capture_interrupt_recovered': False,
        'has_start_local_audio': False, 'has_stop_local_audio': False,
        'capture_start_failed': False, 'no_mic_permission': False,
        'bg_capture_silent': False, 'timeout_exit': False,
        'sample_rate_changed': False,
    }

    bg_enter = None
    bg_return = False
    capture_interrupted = False
    sample_rate_changes = []

    for ev in events:
        v = ev.get('V', 0)
        t = ev.get('T', 0)
        p1 = ev.get('Para1')
        p2 = ev.get('Para2')
        p3 = ev.get('Para3')

        if v == 5003:
            flags['enter_room_ok'] = True
        elif v == 5009:
            flags['first_audio'] = True
        elif v == 7001:
            flags['exit_normal'] = True
        elif v == 2001:
            flags['has_bg_switch'] = True
            if p1 == 1:  # 切后台
                bg_enter = t
                bg_return = False
                findings.append(f'APP切后台({_ms_to_time(t)})')
            elif p1 == 0:  # 回前台
                bg_return = True
                if bg_enter:
                    bg_dur = (t - bg_enter) / 1000
                    findings.append(f'APP回前台({_ms_to_time(t)},后台{bg_dur:.0f}秒)')
                bg_enter = None
        elif v == 3001:
            if p1 == 1:  # 停止采集
                flags['has_mute'] = True
                findings.append(f'停止采集音频({_ms_to_time(t)})')
        elif v == 3011:
            # 本地静音/mute
            if p1 == 1:
                flags['has_mute'] = True
                findings.append(f'mute静音({_ms_to_time(t)})')
            elif p1 == 0:
                findings.append(f'取消mute({_ms_to_time(t)})')
        elif v == 3005:
            flags['has_capture_interrupt'] = True
            capture_interrupted = True
            findings.append(f'采集打断开始({_ms_to_time(t)})')
        elif v == 3006:
            flags['capture_interrupt_recovered'] = True
            capture_interrupted = False
            findings.append(f'采集打断恢复({_ms_to_time(t)})')
        elif v == 3013:
            flags['capture_start_failed'] = True
            findings.append(f'采集启动失败({_ms_to_time(t)})')
        elif v == 3014:
            if p1 == 0:  # 无权限
                flags['no_mic_permission'] = True
                findings.append(f'无麦克风权限({_ms_to_time(t)})')
        elif v == 6001:
            flags['has_start_local_audio'] = True
        elif v == 6002:
            flags['has_stop_local_audio'] = True
        elif v == 3008:
            sample_rate_changes.append({'t': t, 'rate': p1, 'ch': p2})

        # 超时退房检测 (7001 with Para1=2 or specific timeout indicators)
        if v == 7001 and p1 == 2:
            flags['timeout_exit'] = True
            findings.append(f'超时退房({_ms_to_time(t)})')

    # 采样率变化
    if len(sample_rate_changes) >= 2:
        for i in range(1, len(sample_rate_changes)):
            r0 = sample_rate_changes[i - 1]
            r1 = sample_rate_changes[i]
            if r0['rate'] != r1['rate']:
                flags['sample_rate_changed'] = True
                findings.append(
                    f"采样率变更{r0['rate']}→{r1['rate']}Hz"
                )
                break

    # 切后台后未回来
    if bg_enter and not bg_return:
        flags['bg_no_return'] = True
        if flags['exit_normal']:
            findings.append('在后台状态下退出通话')

    # 后台采集无声：切后台 + 未调用startLocalAudio恢复
    if flags['has_bg_switch'] and not flags['has_start_local_audio']:
        flags['bg_capture_silent'] = True

    if not flags['enter_room_ok']:
        findings.append('未成功进入房间')
    if not flags['first_audio']:
        findings.append('未发送首帧音频')

    # 采集打断未恢复
    if flags['has_capture_interrupt'] and not flags['capture_interrupt_recovered']:
        findings.append('采集打断未恢复')

    return {'findings': findings, 'flags': flags}


def _analyze_audio_metrics(cap_energy, play_energy):
    """分析音频采集/播放音量数据
    
    返回 dict:
      - findings: list[str]
      - cap_stats: dict  发送端采集统计
      - play_stats: dict  接收端播放统计
    """
    findings = []
    cap_stats = {'zero_pct': 0, 'avg': 0, 'total': 0, 'silent': False, 'weak': False}
    play_stats = {'zero_pct': 0, 'avg': 0, 'total': 0, 'silent': False, 'weak': False}

    if cap_energy:
        vals = [c['V'] for c in cap_energy if c.get('V') is not None]
        if vals:
            cap_stats['total'] = len(vals)
            zeros = sum(1 for v in vals if v == 0)
            cap_stats['zero_pct'] = zeros / len(vals) * 100
            cap_stats['avg'] = sum(vals) / len(vals)
            if cap_stats['zero_pct'] > 80:
                cap_stats['silent'] = True
                findings.append(f'采集无声(音量{cap_stats["zero_pct"]:.0f}%为零)')
            elif cap_stats['zero_pct'] > 50:
                findings.append(f'采集间歇性静音({cap_stats["zero_pct"]:.0f}%为零)')
            elif cap_stats['avg'] < 200 and cap_stats['avg'] > 0:
                cap_stats['weak'] = True
                findings.append(f'采集弱音(平均音量{cap_stats["avg"]:.0f})')

    if play_energy:
        vals = [c['V'] for c in play_energy if c.get('V') is not None]
        if vals:
            play_stats['total'] = len(vals)
            zeros = sum(1 for v in vals if v == 0)
            play_stats['zero_pct'] = zeros / len(vals) * 100
            play_stats['avg'] = sum(vals) / len(vals)
            if play_stats['zero_pct'] > 80:
                play_stats['silent'] = True
                findings.append(f'播放无声(音量{play_stats["zero_pct"]:.0f}%为零)')
            elif play_stats['zero_pct'] > 50:
                findings.append(f'播放间歇性静音({play_stats["zero_pct"]:.0f}%为零)')

    return {'findings': findings, 'cap_stats': cap_stats, 'play_stats': play_stats}


def _ms_to_time(ts_ms):
    """毫秒级时间戳转 HH:MM:SS"""
    if ts_ms > 1e12:
        return datetime.fromtimestamp(ts_ms / 1000).strftime('%H:%M:%S')
    return datetime.fromtimestamp(ts_ms).strftime('%H:%M:%S')


# ─── 确定对端系统平台 ──────────────────────────────────────
def _detect_peer_platform(reporter_sys, reporter_model, users_info,
                          send_user, recv_user):
    """推断对端(上行端/发送端)的平台: 'android' / 'iOS' / '未知'
    
    优先使用 getUserInfo 返回的 Os/DeviceType 信息。
    注意：不能从反馈者系统反推，因为两端可能都是同一平台（如都是iOS）。
    """
    if users_info and send_user:
        u = users_info.get(send_user, {})
        os_name = str(u.get('Os', '') or '').lower()
        dev = str(u.get('DeviceType', '') or '').lower()
        if 'ios' in os_name or 'iphone' in dev or 'ipad' in dev:
            return 'iOS'
        if 'android' in os_name:
            return 'android'
        # 通过设备型号推断
        if dev:
            # iPhone/iPad 设备名通常以 iphone/ipad 开头
            if any(k in dev for k in ['iphone', 'ipad', 'ipod']):
                return 'iOS'
            # 其他设备名（如 PKD130, V2410A 等）通常是安卓
            return 'android'
    # 无法确定时返回未知，不做反推
    return '未知'


def _detect_reporter_platform(system, model):
    """判断反馈者(接收端)的平台"""
    if system:
        s = str(system).lower()
        if 'ios' in s or 'iphone' in s:
            return 'iOS'
        if '安卓' in s or 'android' in s:
            return 'android'
    if model:
        m = str(model).lower()
        if 'iphone' in m or 'ipad' in m:
            return 'iOS'
        return 'android'
    return '未知'


# ─── 生成排查结论 (v4: 匹配老板简洁风格) ─────────────────
def generate_conclusion(feedback_text, channel_type, phone_model,
                        system, version, network, is_speaker,
                        headphone_type, room_info=None,
                        deep_analysis=None):
    """生成简洁排查结论 + 问题标签
    返回: (conclusion: str, tag: str or None)
    """
    fb = parse_feedback_text(feedback_text)
    problem_type = fb.get('problem_type', '未知')
    reporter_platform = _detect_reporter_platform(system, phone_model)

    if not deep_analysis:
        if not room_info:
            return '未查到通话记录', None
        return '仅有基础信息，需深度分析', None

    users = deep_analysis.get('users', {})
    send_user = deep_analysis.get('send_user', '')
    recv_user = deep_analysis.get('recv_user', '')

    sender_result = deep_analysis.get('sender_analysis', {})
    sender_flags = sender_result.get('flags', {})
    receiver_result = deep_analysis.get('receiver_analysis', {})
    receiver_flags = receiver_result.get('flags', {})

    audio_result = deep_analysis.get('audio_analysis', {})
    cap_stats = audio_result.get('cap_stats', {})
    play_stats = audio_result.get('play_stats', {})

    peer_platform = _detect_peer_platform(
        system, phone_model, users, send_user, recv_user
    )
    user_count = room_info.get('user_count', 0) if room_info else 0
    
    # 发送端是否有用户信息（没有 → 可能没真正进房）
    sender_has_info = bool(users.get(send_user))

    return _infer_conclusion_and_tag(
        problem_type, reporter_platform, peer_platform,
        sender_flags, receiver_flags,
        cap_stats, play_stats, user_count, room_info,
        sender_has_info=sender_has_info
    )


def _infer_conclusion_and_tag(problem_type, reporter_platform, peer_platform,
                               sender_flags, receiver_flags,
                               cap_stats, play_stats,
                               user_count, room_info,
                               sender_has_info=True):
    """推断结论和问题标签。返回 (conclusion, tag)
    
    排查优先级（学习老板人工排查逻辑 v5）：
    1. 房间只有一个人 → 疑似没讲话（老板风格）
    2. 超时退房（发送端）→ 看是否同时有mute
    3. 无麦克风权限
    4. 采集启动失败（看是否有电话抢占）
    5. 采集无声（音量数据优先，关联事件辅助）
    6. 采集弱音
    7. mute静音（需要确认是持续性mute，不是短暂的；同时看音量）
    8. 后台采集问题（区分 android后台采集无声 vs iOS未调startLocalAudio）
    9. 采集打断
    10. 回声/杂音
    11. 疑似没讲话（采集音量低但无明确异常）
    12. 数据正常
    
    关键改进点（对齐老板逻辑）：
    - 采样率变更不作为独立根因，仅在回声场景使用
    - 音量数据优先于事件码判断
    - 切后台判定更严格：需结合音量确认
    - mute判定更严格：需持续性mute才判定
    - 接收端异常用自然语言描述，不暴露内部flag
    - "数据正常"不加平台对比
    - 使用"安卓"而非"android"用于"安卓采集无声/弱音"
    """
    sf = sender_flags or {}
    rf = receiver_flags or {}

    # 辅助函数：用"安卓"还是"android"——老板习惯：
    # "android上行端 mute 静音"、"android 后台采集无声" 用 android
    # "安卓采集无声"、"安卓采集弱音？" 用 安卓
    def peer_label():
        """上行端/发送端平台标签"""
        return peer_platform if peer_platform != '未知' else 'android'

    def reporter_label():
        """接收端/反馈端平台标签"""
        return reporter_platform if reporter_platform != '未知' else 'iOS'

    def peer_cn():
        """中文化平台名（安卓采集无声/弱音等场景）"""
        p = peer_label()
        return '安卓' if p.lower() == 'android' else p

    def reporter_cn():
        p = reporter_label()
        return '安卓' if p.lower() == 'android' else p

    # ===== 1. 房间只有一个人 / 发送端无信息 =====
    if user_count == 1:
        return '房内只有一个人', '房间只有一个人'
    # 发送端查不到用户信息（没真正进房）+ 没有采集音频 → 等效"房内只有一个人"
    if not sender_has_info and cap_stats.get('total', 0) == 0:
        return '房内只有一个人', '房间只有一个人'

    # ===== 2. 超时退房 =====
    if sf.get('timeout_exit'):
        if sf.get('has_mute'):
            return '超时退房+mute静音', '断连超时退房'
        return f'{peer_label()} 上行端超时退房', '断连超时退房'
    if rf.get('timeout_exit'):
        return '超时退房', '断连超时退房'
    # 发送端未成功进房+无首帧音频 → 相当于超时退房
    if sf and not sf.get('enter_room_ok') and not sf.get('first_audio'):
        if sf.get('exit_normal'):
            return f'{peer_label()} 上行端超时退房', '断连超时退房'

    # ===== 3. 无麦克风权限 =====
    if sf.get('no_mic_permission'):
        p = peer_label()
        label = 'iPhone' if p == 'iOS' else p
        return f'{label} 无采集权限', '没有麦克风权限'
    if rf.get('no_mic_permission'):
        p = reporter_label()
        label = 'iPhone' if p == 'iOS' else p
        return f'{label} 无采集权限', '没有麦克风权限'
    # 接收端未发送首帧音频+进房成功 → 可能无采集权限
    if rf and rf.get('enter_room_ok') and not rf.get('first_audio'):
        p = reporter_label()
        label = 'iPhone' if p == 'iOS' else p
        # 如果有采集打断，可能是电话抢占
        if rf.get('capture_start_failed') or rf.get('has_capture_interrupt'):
            return f'{label} 无采集权限', '没有麦克风权限'

    # ===== 4. 采集启动失败 =====
    if sf.get('capture_start_failed'):
        if sf.get('has_capture_interrupt'):
            p = peer_label()
            label = 'iOS' if p == 'iOS' else p
            return f'{label}采集启动失败，\n疑似有系统电话抢占', None
        return f'{peer_label()} 启动采集失败', None
    if rf.get('capture_start_failed'):
        if rf.get('has_capture_interrupt'):
            p = reporter_label()
            label = 'iOS' if p == 'iOS' else p
            return f'{label}采集启动失败，\n疑似有系统电话抢占', None
        return f'{reporter_label()} 启动采集失败', None

    # ===== 5. 采集无声（音量数据优先）=====
    # 老板很重视音量：80%以上为零 → 采集无声
    if cap_stats.get('silent'):
        # 发送端采集无声 → 看原因
        if sf.get('has_bg_switch') and not sf.get('has_start_local_audio'):
            # 切后台+没调startLocalAudio → "android 后台采集无声" 或 "iOS 切后台未执行startLocalAudio"
            p = peer_label()
            if p.lower() == 'android' or p == '未知':
                return 'android 后台采集无声', '安卓后台采集无声'
            return 'iOS 切后台未执行 startLocalAudio', '未调用startlocalaudio'
        if sf.get('has_bg_switch'):
            p = peer_label()
            if p.lower() == 'android' or p == '未知':
                return 'android 后台采集无声', '安卓后台采集无声'
            return f'{p} 后台采集无声', '安卓后台采集无声'
        # 纯采集无声（无后台切换）→ 用"安卓"/"iOS"
        return f'{peer_cn()}采集无声', None

    # ===== 6. 采集弱音 =====
    if cap_stats.get('weak'):
        p = peer_cn()
        avg = cap_stats.get('avg', 0)
        # 老板风格："安卓采集弱音？"、"iOS端采集声音很小 无异常 怀疑当时没说话"
        if p == 'iOS':
            return f'iOS端采集声音很小 无异常 怀疑当时没说话', None
        return f'{p}采集弱音？', None

    # ===== 7. mute 静音 =====
    # 判定规则（从API调查数据学习）：
    # - 采集音量avg > 2000且zero_pct低 → mute短暂，不是主因（跳过）
    # - 采集音量avg < 500 → mute+疑似没讲话
    # - 其他 → android/iOS上行端 mute 静音
    if sf.get('has_mute'):
        cap_avg = cap_stats.get('avg', 0)
        cap_total = cap_stats.get('total', 0)
        cap_zero = cap_stats.get('zero_pct', 0)
        
        if sf.get('capture_start_failed'):
            pass  # 采集启动失败优先
        elif cap_total > 0 and cap_avg > 2000 and cap_zero < 20:
            # 采集音量正常 → mute只是短暂的，跳过
            pass
        elif cap_total > 0 and cap_avg < 500:
            # 音量低 → mute+疑似没讲话
            return 'mute静音+疑似没有人讲话', 'mute操作'
        elif cap_total == 0:
            # 无采集数据 → mute静音
            return f'{peer_label()}上行端 mute 静音', 'mute操作'
        else:
            # 有采集数据且音量不太高 → mute是主因
            return f'{peer_label()}上行端 mute 静音', 'mute操作'

    # ===== 8. 后台采集问题 =====
    # 区分 android后台采集无声 vs iOS切后台未startLocalAudio
    if sf.get('has_bg_switch'):
        p = peer_label()
        cap_avg = cap_stats.get('avg', 0)
        cap_total = cap_stats.get('total', 0)
        cap_zero = cap_stats.get('zero_pct', 0)

        # 如果采集音量完全正常（不是无声也不是弱音），切后台可能不是主因
        if cap_total > 0 and cap_avg > 1000 and cap_zero < 30:
            pass  # 跳过，让后续逻辑处理
        elif not sf.get('has_start_local_audio'):
            if p.lower() == 'android' or p == '未知':
                # android 切后台 → "android 后台采集无声"
                return f'android 后台采集无声', '安卓后台采集无声'
            else:
                # iOS 切后台 → "iOS 切后台未执行 startLocalAudio"
                return f'iOS 切后台未执行 startLocalAudio', '未调用startlocalaudio'
        elif sf.get('bg_no_return'):
            return f'{p} 后台采集无声', '安卓后台采集无声'

    # ===== 9. 采集打断 =====
    if sf.get('has_capture_interrupt'):
        recovered = sf.get('capture_interrupt_recovered', False)
        p = peer_label()
        label = 'iphone' if p == 'iOS' else p
        if not recovered:
            return f'{label}上行端采播打断', '正常采集打断事件'
        return f'{label}采集打断(已恢复)', '正常采集打断事件'
    if rf.get('has_capture_interrupt'):
        if not rf.get('capture_interrupt_recovered', False):
            p = reporter_label()
            label = 'iPhone' if p == 'iOS' else p
            return f'{label}采集打断', '正常采集打断事件'

    # ===== 10. 回声/杂音 =====
    if problem_type == '有回音/杂音':
        if sf.get('sample_rate_changed') or rf.get('sample_rate_changed'):
            return '采样率变更导致回声', '回声问题'
        return 'iOS 漏回声？', '回声问题'

    # ===== 11. 疑似没讲话 =====
    # 采集音量低但无明确异常事件 → 疑似没讲话
    # 阈值放宽：avg < 1500 就算低（老板判"疑似没讲话"的范围更宽）
    cap_avg = cap_stats.get('avg', 0)
    cap_total = cap_stats.get('total', 0)
    if cap_total > 0 and cap_avg < 1500 and cap_avg > 0:
        return '疑似没讲话', None
    # 没有采集数据（total=0）但有播放数据 → 也可能是没讲话
    if cap_total == 0 and play_stats.get('total', 0) > 0:
        return '疑似没讲话', None

    # ===== 12. 接收端后台切换问题 =====
    if rf.get('has_bg_switch'):
        if rf.get('has_capture_interrupt') and not rf.get('capture_interrupt_recovered'):
            p = reporter_label()
            label = 'iPhone' if p == 'iOS' else p
            return f'{label}采集打断', '正常采集打断事件'
        if not rf.get('has_start_local_audio'):
            # 接收端切后台 → 但这一般不影响上行端
            pass

    # ===== 13. 数据正常 / 疑似没讲话 =====
    # 如果采集和播放数据都没有（total=0），也可能是没讲话
    if cap_stats.get('total', 0) == 0 and play_stats.get('total', 0) == 0:
        return '疑似没讲话', None
    # 老板风格：直接写 "数据正常"
    return '数据正常', None


def _has_any_anomaly(flags):
    """检查是否有任何异常标志
    
    注意：sample_rate_changed 不作为独立异常（老板逻辑），仅在回声场景使用
    """
    if not flags:
        return False
    for k in ('has_bg_switch', 'bg_no_return', 'has_mute',
              'has_capture_interrupt', 'capture_start_failed',
              'no_mic_permission', 'timeout_exit',
              'has_stop_local_audio'):
        if flags.get(k):
            return True
    if not flags.get('enter_room_ok'):
        return True
    if not flags.get('first_audio'):
        return True
    return False


# ─── Playwright 登录 + API 调用 ──────────────────────────
def ensure_login_and_fetch(rows_to_query, headless=False):
    """通过 Playwright 持久化上下文确保已登录，然后批量调 API"""
    if not rows_to_query:
        return {}

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("错误: playwright 未安装。请运行: pip install playwright && playwright install chromium", flush=True)
        return {}

    results = {}

    with sync_playwright() as p:
        print(f"启动浏览器 (持久化目录: {PW_USER_DATA_DIR}, headless={headless})", flush=True)

        # 清理可能残留的锁文件
        for lock in ('SingletonLock', 'SingletonSocket', 'SingletonCookie'):
            lock_path = os.path.join(PW_USER_DATA_DIR, lock)
            if os.path.exists(lock_path):
                try:
                    os.remove(lock_path)
                except OSError:
                    pass

        browser = p.chromium.launch_persistent_context(
            PW_USER_DATA_DIR,
            headless=headless,
            args=['--disable-blink-features=AutomationControlled'],
            viewport={'width': 1280, 'height': 800}
        )
        page = browser.pages[0] if browser.pages else browser.new_page()

        # Step 1: 检查是否需要登录
        print("检查登录状态...", flush=True)
        page.goto(f'https://{TRTC_MONITOR_HOST}',
                  timeout=30000, wait_until='domcontentloaded')
        time.sleep(2)

        current_url = page.url
        if OA_LOGIN_HOST in current_url:
            if headless:
                print("❌ headless 模式下登录态已失效，无法交互登录。", flush=True)
                print("   请先用非 headless 模式运行一次完成 OA 登录：", flush=True)
                print("   python3 gen_dashboard.py <input> <output> --deep", flush=True)
                browser.close()
                return {}
            print("\n" + "=" * 60, flush=True)
            print("⚠️  需要 OA 登录！", flush=True)
            print("请在弹出的浏览器窗口中完成 iOA 验证", flush=True)
            print("（点击「发起验证」→ 手机 iOA 确认）", flush=True)
            print("=" * 60 + "\n", flush=True)

            # 等待用户完成登录 (最多等5分钟)
            try:
                page.wait_for_url(
                    f'https://{TRTC_MONITOR_HOST}/**',
                    timeout=300000  # 5分钟
                )
                print("✅ 登录成功！Cookie 已保存，下次无需重新登录。\n", flush=True)
                time.sleep(2)
            except Exception:
                print("❌ 登录超时（5分钟），请重新运行。", flush=True)
                browser.close()
                return {}
        else:
            print("✅ 已有登录态，直接使用。\n", flush=True)

        # Step 2: 批量调用 API
        total = len(rows_to_query)
        deep = any(item.get('deep') for item in rows_to_query)
        print(f"开始查询 {total} 条通话记录{'(含深度分析)' if deep else ''}...", flush=True)

        def _fetch(url):
            """fetch API 并返回 JSON"""
            resp = page.evaluate("""
                async (url) => {
                    const resp = await fetch(url, { credentials: 'include' });
                    return await resp.text();
                }
            """, url)
            return json.loads(resp)

        for idx, item in enumerate(rows_to_query, 1):
            row_idx = item['row_idx']
            api_url = item['api_url']
            sdkappid = item['sdkappid']
            channel_id = item['channel_id']
            do_deep = item.get('deep', False)
            send_user_id = item.get('send_user', '')
            recv_user_id = item.get('recv_user', '')

            try:
                data = _fetch(api_url)
                room_list = _extract_room_list(data)

                if room_list:
                    room = room_list[0]
                    room_info = _extract_room_info(room, channel_id, sdkappid)
                    if room_info.get('comm_id'):
                        detail_url = build_detail_url(**room_info['url_params'])
                        result_entry = {
                            'detail_url': detail_url,
                            'room_info': room_info,
                        }

                        # ─── 深度分析 ───
                        if do_deep and room_info.get('comm_id'):
                            deep_data = _fetch_deep_analysis(
                                _fetch, room_info, sdkappid,
                                send_user_id, recv_user_id
                            )
                            result_entry['deep_analysis'] = deep_data
                            n_findings = (
                                len(deep_data.get('sender_analysis', {}).get('findings', []))
                                + len(deep_data.get('receiver_analysis', {}).get('findings', []))
                                + len(deep_data.get('audio_analysis', {}).get('findings', []))
                            )
                            print(f"  [{idx}/{total}] 行{row_idx}: ✅ CommId={room_info['comm_id']} 深度分析发现{n_findings}项", flush=True)
                        else:
                            print(f"  [{idx}/{total}] 行{row_idx}: ✅ CommId={room_info['comm_id']}", flush=True)

                        results[row_idx] = result_entry
                    else:
                        print(f"  [{idx}/{total}] 行{row_idx}: ⚠️ 无CommId", flush=True)
                else:
                    print(f"  [{idx}/{total}] 行{row_idx}: ⚠️ 无通话记录", flush=True)
                    results[row_idx] = {
                        'detail_url': None,
                        'room_info': None,
                    }

            except Exception as e:
                print(f"  [{idx}/{total}] 行{row_idx}: ❌ {e}", flush=True)
                results[row_idx] = {
                    'detail_url': None,
                    'room_info': None,
                }

            if idx < total:
                time.sleep(0.3)

        print(f"\n查询完成。成功: {sum(1 for v in results.values() if v.get('detail_url'))}/{total}", flush=True)
        browser.close()

    return results


def _fetch_deep_analysis(fetch_fn, room_info, sdkappid,
                         send_user_id, recv_user_id):
    """获取深度分析数据: getUserInfo + detail_event + 音频指标"""
    comm_id = room_info['comm_id']
    create_ts = room_info['create_time']
    destroy_ts = room_info['destroy_time']
    room_num = room_info['url_params']['room_num']

    result = {
        'users': {},
        'sender_analysis': {'findings': [], 'flags': {}},
        'receiver_analysis': {'findings': [], 'flags': {}},
        'audio_analysis': {'findings': [], 'cap_stats': {}, 'play_stats': {}},
        'send_user': send_user_id,
        'recv_user': recv_user_id,
    }

    # 1. 获取两端用户信息 + TinyId
    for uid in [send_user_id, recv_user_id]:
        if not uid:
            continue
        try:
            url = (
                f"https://avmonitor.trtc.woa.com/getUserInfo"
                f"?SdkAppId={sdkappid}"
                f"&CommId={comm_id}"
                f"&StartTs={create_ts}"
                f"&EndTs={destroy_ts}"
                f"&UserId={uid}"
            )
            data = fetch_fn(url)
            user_list = data.get('Response', {}).get('UserList', [])
            if user_list:
                result['users'][uid] = user_list[0]
        except Exception:
            pass
        time.sleep(0.1)

    # 2. 获取事件列表 (detail_event)
    for uid, role in [(send_user_id, 'sender'), (recv_user_id, 'receiver')]:
        if not uid:
            continue
        tiny_id = result['users'].get(uid, {}).get('TinyId', '')
        if not tiny_id:
            continue
        try:
            url = (
                f"https://avmonitor.trtc.woa.com/getElasticSearchData"
                f"?StartTs={create_ts}"
                f"&EndTs={destroy_ts}"
                f"&CommId={comm_id}"
                f"&UserId={uid}"
                f"&TinyId={tiny_id}"
                f"&RoomNum={room_num}"
                f"&IndexType=event"
                f"&DataType=detail_event"
            )
            data = fetch_fn(url)
            resp_data = data.get('Response', {}).get('Data', [])
            if resp_data:
                events = resp_data[0].get('Content', [])
                analysis = _analyze_events(events)
                result[f'{role}_analysis'] = analysis
        except Exception:
            pass
        time.sleep(0.1)

    # 3. 获取音频采集音量 (发送端上行)
    send_tiny = result['users'].get(send_user_id, {}).get('TinyId', '')
    recv_tiny = result['users'].get(recv_user_id, {}).get('TinyId', '')

    cap_energy = []
    if send_user_id and send_tiny:
        try:
            url = (
                f"https://avmonitor.trtc.woa.com/getElasticSearchData"
                f"?StartTs={create_ts}&EndTs={destroy_ts}"
                f"&CommId={comm_id}"
                f"&UserId={send_user_id}"
                f"&TinyId={send_tiny}"
                f"&RoomNum={room_num}"
                f"&IndexType=up&DataType=aCapEnergy"
            )
            data = fetch_fn(url)
            resp_data = data.get('Response', {}).get('Data', [])
            if resp_data:
                cap_energy = resp_data[0].get('Content', [])
        except Exception:
            pass

    # 4. 获取音频播放音量 (接收端下行)
    play_energy = []
    if recv_user_id and recv_tiny:
        try:
            url = (
                f"https://avmonitor.trtc.woa.com/getElasticSearchData"
                f"?StartTs={create_ts}&EndTs={destroy_ts}"
                f"&CommId={comm_id}"
                f"&UserId={recv_user_id}"
                f"&TinyId={recv_tiny}"
                f"&RoomNum={room_num}"
                f"&IndexType=down&DataType=aPlayEnergy"
            )
            data = fetch_fn(url)
            resp_data = data.get('Response', {}).get('Data', [])
            if resp_data:
                play_energy = resp_data[0].get('Content', [])
        except Exception:
            pass

    result['audio_analysis'] = _analyze_audio_metrics(cap_energy, play_energy)

    return result


def _extract_room_list(data):
    """从 API 响应中提取房间列表，兼容多种响应格式"""
    # 常见路径（按优先级排列）
    for path in [
        lambda d: d.get('Response', {}).get('RoomList'),   # 实际格式
        lambda d: d.get('Data', {}).get('RoomList'),
        lambda d: d.get('data', {}).get('roomList'),
        lambda d: d.get('Data', {}).get('roomList'),
        lambda d: d.get('response', {}).get('RoomList'),
        lambda d: d.get('Data') if isinstance(d.get('Data'), list) else None,
    ]:
        try:
            result = path(data)
            if result and isinstance(result, list):
                return result
        except (AttributeError, TypeError):
            continue

    # 遍历 Response / Data 下所有 list 类型字段
    for key in ('Response', 'Data', 'response', 'data'):
        data_obj = data.get(key)
        if isinstance(data_obj, dict):
            for val in data_obj.values():
                if isinstance(val, list) and val:
                    return val

    return []


def _extract_room_info(room, channel_id, sdkappid):
    """从单条房间记录中提取信息，兼容多种字段命名"""
    def get(room, *keys):
        for k in keys:
            v = room.get(k)
            if v is not None:
                return v
        return None

    comm_id = get(room, 'CommId', 'commId', 'comm_id') or ''
    room_num = get(room, 'RoomNum', 'roomNum', 'room_num') or ''
    room_str = get(room, 'RoomStr', 'roomStr', 'room_str') or channel_id
    create_time = get(room, 'CreateTs', 'CreateTime', 'createTime', 'create_time') or 0
    destroy_time = get(room, 'DestroyTs', 'DestroyTime', 'destroyTime', 'destroy_time') or 0
    duration = get(room, 'Duration', 'duration') or 0
    finished = get(room, 'Finished', 'finished', 'IsFinished')
    if finished is None:
        finished = True
    user_count = get(room, 'UserNum', 'UserCount', 'userCount', 'user_count') or 0

    return {
        'comm_id': comm_id,
        'duration': duration,
        'finished': finished,
        'create_time': create_time,
        'destroy_time': destroy_time,
        'user_count': user_count,
        'url_params': {
            'comm_id': comm_id,
            'room_num': room_num,
            'room_str': room_str if room_str else channel_id,
            'create_time': create_time,
            'destroy_time': destroy_time,
            'duration': duration,
            'finished': finished,
            'sdkappid': sdkappid,
        }
    }


# ─── 主流程 ──────────────────────────────────────────────
def process(input_path, output_path, detail=False, analyze=False,
            deep_analyze=False, headless=False):
    wb = load_workbook(input_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    col_channel = find_col(header, '反馈渠道')
    col_channel_id = find_col(header, 'channelId')
    col_feedback_time = find_col(header, '反馈时间')
    col_feedback_text = find_col(header, '反馈文本')
    col_phone_model = find_col(header, '手机型号')
    col_system = find_col(header, '系统')
    col_version = find_col(header, '版本')
    col_network = find_col(header, 'network')
    col_is_speaker = find_col(header, '是否为扬声器')
    col_headphone = find_col(header, '耳机类型')

    if not all([col_channel, col_channel_id, col_feedback_time]):
        print(f"错误：找不到必需列。"
              f"反馈渠道={col_channel}, "
              f"channelId={col_channel_id}, "
              f"反馈时间={col_feedback_time}")
        sys.exit(1)

    # 设置表头 (不覆盖已有表头)
    if not ws.cell(row=1, column=TAG_COL).value:
        ws.cell(row=1, column=TAG_COL).value = '问题标签'
    if not ws.cell(row=1, column=CONCLUSION_COL).value:
        ws.cell(row=1, column=CONCLUSION_COL).value = '排查结论'
    if not ws.cell(row=1, column=SEARCH_COL).value:
        ws.cell(row=1, column=SEARCH_COL).value = '仪表盘'
    if detail and not ws.cell(row=1, column=DETAIL_COL).value:
        ws.cell(row=1, column=DETAIL_COL).value = '通话详情'

    rows_to_query = []

    # Step 1: 遍历所有行，收集需要查详情的行
    for row_idx in range(2, ws.max_row + 1):
        channel_type = ws.cell(row=row_idx, column=col_channel).value
        channel_id = ws.cell(row=row_idx, column=col_channel_id).value
        feedback_time = ws.cell(row=row_idx, column=col_feedback_time).value

        if not channel_type or not channel_id or not feedback_time:
            continue

        sdkappid = SDKAPPID_MAP.get(str(channel_type).strip())
        if not sdkappid:
            continue

        ts = to_timestamp(feedback_time)
        if ts is None:
            continue

        cid = str(channel_id).strip()
        start_ts = ts - TIME_RANGE
        end_ts = ts + TIME_RANGE

        # 收集需要查询详情的行 (U列和V列都填通话详情URL)
        if detail:
            # 只要 U列 或 V列 任一缺 URL，就需要查询
            existing_search = ws.cell(row=row_idx, column=SEARCH_COL).value
            existing_detail = ws.cell(row=row_idx, column=DETAIL_COL).value
            search_has_url = existing_search and str(existing_search).strip().startswith('http')
            detail_has_url = existing_detail and str(existing_detail).strip().startswith('http')
            if not search_has_url or not detail_has_url:
                api_url = build_roomlist_api_url(
                    sdkappid, cid, start_ts, end_ts
                )
                # 从反馈文本中提取 peerUIDs
                fb_text = str(ws.cell(row=row_idx,
                              column=col_feedback_text).value or '')
                peer_m = re.search(r'peerUIDs:([^\n]+)', fb_text)
                peer_uids = []
                if peer_m:
                    peer_uids = [u.strip() for u in
                                 peer_m.group(1).split(',') if u.strip()]
                # 提取呼叫方/被呼叫方 uid (D/E列,加密)
                col_caller = find_col(header, '呼叫方uid')
                col_callee = find_col(header, '被呼叫方uid')
                caller_uid = str(ws.cell(row=row_idx,
                                 column=col_caller).value or '') if col_caller else ''
                callee_uid = str(ws.cell(row=row_idx,
                                 column=col_callee).value or '') if col_callee else ''

                # 确定 send/recv: 反馈方是接收端，对方是发送端
                send_user = ''
                recv_user = ''
                if len(peer_uids) >= 2:
                    # peerUIDs 中的第一个通常是反馈方自己
                    recv_user = peer_uids[0]  # 反馈方
                    send_user = peer_uids[1]  # 对方
                elif caller_uid and callee_uid:
                    send_user = caller_uid
                    recv_user = callee_uid

                rows_to_query.append({
                    'row_idx': row_idx,
                    'api_url': api_url,
                    'sdkappid': sdkappid,
                    'channel_id': cid,
                    'deep': deep_analyze,
                    'send_user': send_user,
                    'recv_user': recv_user,
                })

    # Step 2: 批量获取通话详情
    api_results = {}
    filled_search = 0
    filled_detail = 0
    if detail and rows_to_query:
        print(f"\n需要查询 {len(rows_to_query)} 条通话详情...", flush=True)
        api_results = ensure_login_and_fetch(rows_to_query, headless=headless)

        for row_idx, info in api_results.items():
            detail_url = info.get('detail_url')
            if detail_url:
                # U列（仪表盘）也写入通话详情 URL
                cell_search = ws.cell(row=row_idx, column=SEARCH_COL)
                if not (cell_search.value and str(cell_search.value).strip().startswith('http')):
                    cell_search.value = detail_url
                    cell_search.hyperlink = detail_url
                    cell_search.font = LINK_FONT
                    filled_search += 1
                # V列（通话详情）
                cell_detail = ws.cell(row=row_idx, column=DETAIL_COL)
                if not (cell_detail.value and str(cell_detail.value).strip().startswith('http')):
                    cell_detail.value = detail_url
                    cell_detail.hyperlink = detail_url
                    cell_detail.font = LINK_FONT
                    filled_detail += 1

        # 同步：确保所有行的 U列 与 V列 保持一致（通话详情URL）
        synced = 0
        for row_idx in range(2, ws.max_row + 1):
            v_val = ws.cell(row=row_idx, column=DETAIL_COL).value
            u_val = ws.cell(row=row_idx, column=SEARCH_COL).value
            if v_val and str(v_val).strip().startswith('http'):
                # V列有URL但U列没有或者是旧的搜索页URL -> 同步
                if not u_val or str(u_val).strip() != str(v_val).strip():
                    cell_u = ws.cell(row=row_idx, column=SEARCH_COL)
                    cell_u.value = v_val
                    cell_u.hyperlink = str(v_val)
                    cell_u.font = LINK_FONT
                    synced += 1
                    if u_val and 'call-details' not in str(u_val):
                        filled_search += 1  # 旧搜索页被替换
        if synced:
            print(f"  同步 V列→U列: {synced} 条", flush=True)

    # Step 3: 生成排查结论
    filled_conclusion = 0
    if analyze:
        print(f"\n生成排查结论{'(深度分析)' if deep_analyze else ''}...", flush=True)
        for row_idx in range(2, ws.max_row + 1):
            # 深度分析模式覆盖已有结论；普通模式跳过
            if not deep_analyze:
                existing_conclusion = ws.cell(
                    row=row_idx, column=CONCLUSION_COL
                ).value
                if existing_conclusion and str(existing_conclusion).strip():
                    continue

            channel_type = ws.cell(row=row_idx, column=col_channel).value
            if not channel_type:
                continue

            feedback_text = ws.cell(row=row_idx, column=col_feedback_text).value if col_feedback_text else None
            phone_model = ws.cell(row=row_idx, column=col_phone_model).value if col_phone_model else None
            system = ws.cell(row=row_idx, column=col_system).value if col_system else None
            version = ws.cell(row=row_idx, column=col_version).value if col_version else None
            network = ws.cell(row=row_idx, column=col_network).value if col_network else None
            is_speaker = ws.cell(row=row_idx, column=col_is_speaker).value if col_is_speaker else None
            headphone = ws.cell(row=row_idx, column=col_headphone).value if col_headphone else None

            room_info = None
            deep_data = None
            if row_idx in api_results:
                room_info = api_results[row_idx].get('room_info')
                deep_data = api_results[row_idx].get('deep_analysis')

            conclusion, tag = generate_conclusion(
                feedback_text, channel_type, phone_model,
                system, version, network, is_speaker,
                headphone, room_info, deep_data
            )

            ws.cell(row=row_idx, column=CONCLUSION_COL).value = conclusion
            if tag:
                ws.cell(row=row_idx, column=TAG_COL).value = tag
            filled_conclusion += 1

    # 调整列宽
    ws.column_dimensions['S'].width = 25   # 问题标签
    ws.column_dimensions['T'].width = 50   # 排查结论(简洁风格)
    ws.column_dimensions['U'].width = 80   # 仪表盘
    if detail:
        ws.column_dimensions['V'].width = 120  # 通话详情

    wb.save(output_path)
    print(f"\n{'=' * 50}", flush=True)
    print(f"✅ 完成！", flush=True)
    if detail:
        print(f"  仪表盘(通话详情) URL: {filled_search} 条  (U列)", flush=True)
        print(f"  通话详情 URL:         {filled_detail} 条  (V列)", flush=True)
        print(f"  (共查询 {len(rows_to_query)} 条)", flush=True)
    if analyze:
        print(f"  排查结论:     {filled_conclusion} 条  (T列)", flush=True)
        print(f"  问题标签:     {filled_conclusion} 条  (S列)", flush=True)
    print(f"  输出文件:     {output_path}", flush=True)
    print(f"{'=' * 50}", flush=True)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Soul 工单仪表盘 URL 自动拼接 + 排查分析'
    )
    parser.add_argument('input', help='输入 Excel 文件路径')
    parser.add_argument('output', help='输出 Excel 文件路径')
    parser.add_argument(
        '--detail', action='store_true',
        help='获取通话详情页 URL (首次需要 OA 登录)'
    )
    parser.add_argument(
        '--analyze', action='store_true',
        help='根据反馈文本+通话详情生成排查结论 (T列)'
    )
    parser.add_argument(
        '--deep', action='store_true',
        help='深度分析: 获取事件列表+音频指标, 推断根因 (需更多API调用)'
    )
    parser.add_argument(
        '--headless', action='store_true',
        help='无头模式运行浏览器(需已有登录态)'
    )
    args = parser.parse_args()

    # --deep 隐含 --analyze 和 --detail
    if args.deep:
        args.analyze = True
        args.detail = True

    # --analyze 隐含 --detail
    if args.analyze:
        args.detail = True

    process(args.input, args.output, args.detail, args.analyze, args.deep,
            getattr(args, 'headless', False))
